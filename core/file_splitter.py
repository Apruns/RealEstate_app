# core/file_splitter.py

import os
import csv
import shutil
from datetime import date
from typing import List, Optional, Dict, Tuple
import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

CHUNK_SIZE = 50000

def _get_headers_and_iter_excel(path: str):
    """Yields (headers, row_iterator) for Excel."""
    if load_workbook is None:
        raise ImportError("openpyxl is required for Excel files.")
    
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.rows
    
    try:
        header_cells = next(rows)
        headers = [c.value for c in header_cells]
    except StopIteration:
        headers = []
    
    # Generator that yields lists of values
    def row_gen():
        for r in rows:
            yield [c.value for c in r]
        wb.close()
            
    return headers, row_gen()

def _get_headers_and_iter_csv(path: str):
    """Yields (headers, row_iterator) for CSV, detecting encoding."""
    encoding = "utf-8"
    # Detect encoding
    for enc in ["utf-8", "cp1255", "latin1"]:
        try:
            with open(path, 'r', encoding=enc) as f:
                next(f)
            encoding = enc
            break
        except:
            continue
            
    f = open(path, 'r', encoding=encoding, newline='')
    reader = csv.reader(f)
    try:
        headers = next(reader)
    except StopIteration:
        headers = []
        
    def row_gen():
        for r in reader:
            yield r
        f.close()
        
    return headers, row_gen()

def split_file_to_chunks(input_path: str, temp_dir: str) -> Tuple[List[str], Dict[str, Any]]:
    """
    Reads an Input File (Excel/CSV) and splits it into multiple CSV chunks on disk.
    Also calculates metadata (like max_scan_date) ON THE FLY to avoid re-reading.
    
    Returns:
        (list_of_chunk_paths, metadata_dict)
    """
    os.makedirs(temp_dir, exist_ok=True)
    ext = os.path.splitext(input_path)[1].lower()
    
    # 1. Get Iterator based on file type
    if ext in [".xls", ".xlsx", ".xlsm"]:
        headers, row_iter = _get_headers_and_iter_excel(input_path)
    else:
        headers, row_iter = _get_headers_and_iter_csv(input_path)
        
    # Prepare Metadata tracking
    scan_date_idx = -1
    # Find scan_date column index (case-insensitive)
    for i, h in enumerate(headers):
        if str(h).lower().strip() == "scan_date":
            scan_date_idx = i
            break
            
    max_scan_date = None
    chunk_paths = []
    
    # 2. Iterate and Split
    current_chunk_idx = 0
    current_rows = []
    
    def save_chunk(rows, idx):
        if not rows: return None
        c_name = f"chunk_{idx}.csv"
        c_path = os.path.join(temp_dir, c_name)
        
        # Write to clean CSV
        with open(c_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers) # Write headers to every chunk
            writer.writerows(rows)
        return c_path

    for row in row_iter:
        # Track Metadata (Max Date)
        if scan_date_idx != -1 and len(row) > scan_date_idx:
            val = row[scan_date_idx]
            if val:
                try:
                    # Fast parse check
                    dt = pd.to_datetime(val, dayfirst=True, errors='coerce')
                    if pd.notna(dt):
                        if max_scan_date is None or dt > max_scan_date:
                            max_scan_date = dt
                except:
                    pass

        current_rows.append(row)
        
        # Dump chunk if full
        if len(current_rows) >= CHUNK_SIZE:
            path = save_chunk(current_rows, current_chunk_idx)
            if path: chunk_paths.append(path)
            current_rows = []
            current_chunk_idx += 1
            
    # Save last chunk
    if current_rows:
        path = save_chunk(current_rows, current_chunk_idx)
        if path: chunk_paths.append(path)

    metadata = {
        "max_scan_date": max_scan_date.date() if max_scan_date else None,
        "original_filename": os.path.basename(input_path)
    }
    
    return chunk_paths, metadata
