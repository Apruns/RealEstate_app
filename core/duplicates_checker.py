# core/duplicates_checker.py

import os
from datetime import date
from typing import Dict, Any, List, Tuple, Optional, Generator
import pandas as pd
import numpy as np

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ----------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------

DUP_KEY_COLUMNS: List[str] = [
    "block_lot",
    "sale_day",
    "declared_profit",
    "sold_part",
    "city",
    "build_year",
    "building_mr",
    "rooms_number",
    "scan_date"
]

CHUNK_SIZE = 50000  # Process 50k rows at a time

# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------

def _detect_encoding(scan_path: str) -> str:
    """Try to detect encoding by reading a tiny bit of the file."""
    for enc in ["utf-8", "cp1255", "latin1"]:
        try:
            pd.read_csv(scan_path, nrows=100, encoding=enc)
            return enc
        except Exception:
            continue
    return "utf-8"  

def _excel_chunk_iterator(excel_path: str, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
    """
    Yields chunks of the Excel file as DataFrames directly from the stream.
    Used for main processing.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required. pip install openpyxl")

    # read_only=True is critical for low memory usage
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    rows_iter = ws.rows
    
    # 1. Get Headers
    try:
        header_row = next(rows_iter)
        headers = [cell.value for cell in header_row]
    except StopIteration:
        wb.close()
        return

    # 2. Accumulate rows and yield chunks
    current_chunk = []
    
    for row in rows_iter:
        # Get values
        values = [cell.value for cell in row]
        current_chunk.append(values)
        
        if len(current_chunk) >= chunk_size:
            yield pd.DataFrame(current_chunk, columns=headers)
            current_chunk = []  # Reset buffer
            
    # Yield remaining
    if current_chunk:
        yield pd.DataFrame(current_chunk, columns=headers)
        
    wb.close()

def _read_and_filter_chunks(scan_path: str) -> Tuple[pd.DataFrame, int]:
    """
    Reads file in chunks.
    REMOVED: Date filtering.
    KEEPS: Only rows where sold_part==1.
    Returns: (filtered_df, total_rows_scanned)
    """
    ext = os.path.splitext(scan_path)[1].lower()
    total_rows = 0
    filtered_chunks = []
    
    # Setup iterator
    if ext == ".csv":
        encoding = _detect_encoding(scan_path)
        iterator = pd.read_csv(scan_path, chunksize=CHUNK_SIZE, encoding=encoding, dtype=str)
    else:
        # Excel Stream Iterator
        iterator = _excel_chunk_iterator(scan_path, CHUNK_SIZE)

    for chunk in iterator:
        total_rows += len(chunk)
        
        # Normalize columns (case insensitive find)
        cols_map = {str(c).lower(): c for c in chunk.columns}
        
        # 1. Filter by sold_part == 1
        # We assume we still want to filter logic validity (sold_part=1)
        # even if we are checking the whole file date-wise.
        if "sold_part" in cols_map:
            real_col = cols_map["sold_part"]
            sp = chunk[real_col].astype(str).str.replace(r"[^0-9\.]", "", regex=True)
            sp_numeric = pd.to_numeric(sp, errors="coerce")
            sold_mask = (sp_numeric == 1)
        else:

            sold_mask = pd.Series([False] * len(chunk), index=chunk.index)

        # Apply mask
        if sold_mask.any():
            # Keep only the rows that passed
            filtered_chunks.append(chunk[sold_mask].copy())

    if not filtered_chunks:
        return pd.DataFrame(columns=DUP_KEY_COLUMNS), total_rows
        
    return pd.concat(filtered_chunks, ignore_index=True), total_rows

# ----------------------------------------------------------------------
# Runner
# ----------------------------------------------------------------------

def run_duplicates_check(
    scan_path: str,
    output_dir: str,
    sample_limit: int = 100,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:

    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Read & Filter in Chunks (Memory Efficient)
    # No longer passing a date, it reads the whole file (filtering only logic like sold_part)
    df_filtered, rows_before = _read_and_filter_chunks(scan_path)
    rows_after_filter = len(df_filtered)

    if df_filtered.empty:
        return {
            "rows_before": rows_before,
            "rows_after_filter": 0,
            "latest_scan_date": "All Dates",
            "duplicate_groups": 0,
            "duplicate_rows": 0,
            "output_filename": None,
        }, []

    # 2. Clean numeric columns for grouping
    # Normalize columns to ensure we find "declared_profit" etc. even if casing differs
    ci_map = {str(c).lower(): c for c in df_filtered.columns}
    
    # Map required keys to actual columns
    actual_group_cols = []
    for key in DUP_KEY_COLUMNS:
        if key.lower() in ci_map:
            actual_group_cols.append(ci_map[key.lower()])
        else:
            # If a key column is missing, create it as empty/0
            col_name = key
            df_filtered[col_name] = "0"
            actual_group_cols.append(col_name)

    # Clean numeric keys
    numeric_keys = ["declared_profit", "sold_part", "build_year", "building_mr", "rooms_number"]
    for key in numeric_keys:
        if key.lower() in ci_map:
            col = ci_map[key.lower()]
            df_filtered[col] = df_filtered[col].astype(str).str.replace(r"[^0-9\.\-]", "", regex=True)
            df_filtered[col] = df_filtered[col].replace("", "0")

    # 3. Group By (Find duplicates)
    dup_groups = (
        df_filtered
        .groupby(actual_group_cols, dropna=False)
        .size()
        .reset_index(name="dup_count")
    )
    dup_groups = dup_groups[dup_groups["dup_count"] > 1]

    if dup_groups.empty:
        return {
            "rows_before": rows_before,
            "rows_after_filter": rows_after_filter,
            "latest_scan_date": "All Dates",
            "duplicate_groups": 0,
            "duplicate_rows": 0,
            "output_filename": None,
        }, []

    # 4. Merge back to get actual rows
    dup_groups["dup_group_id"] = dup_groups.index + 1
    dup_rows = df_filtered.merge(
        dup_groups[actual_group_cols + ["dup_count", "dup_group_id"]],
        on=actual_group_cols,
        how="inner",
    )
    dup_rows = dup_rows.sort_values(by=["dup_group_id", "dup_count"], ascending=[True, False])

    # 5. Export
    base_name = os.path.splitext(os.path.basename(scan_path))[0]
    today_str = date.today().strftime("%Y%m%d")
    output_filename = f"duplicates_ALL_DATES_{base_name}_{today_str}.csv"
    output_path = os.path.join(output_dir, output_filename)
    
    dup_rows.to_csv(output_path, index=False, encoding="utf-8-sig")

    results = {
        "rows_before": rows_before,
        "rows_after_filter": rows_after_filter,
        "latest_scan_date": "All Dates",
        "duplicate_groups": int(dup_groups["dup_group_id"].nunique()),
        "duplicate_rows": int(len(dup_rows)),
        "output_filename": output_filename,
    }

    sample_rows = dup_rows.head(sample_limit).to_dict(orient="records")
    return results, sample_rows