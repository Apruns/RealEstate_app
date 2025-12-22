# core/prepare_yzer.py

import os
from datetime import date
from typing import Dict, Any, List, Generator
import pandas as pd
import numpy as np

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ---------------------------------------------------------
# Config
# ---------------------------------------------------------
PREFERRED_COLUMN_ORDER: List[str] = [
    "id", "code", "start_block", "end_block", "start_lot", "end_lot", "page", "line",
    "block_lot", "sale_day", "declared_profit", "sale_profit", "property_type", "sold_part",
    "full_price", "city", "build_year", "building_mr", "rooms_number", "area", "goch",
    "deal_date", "city2", "street", "home_num", "entrance_num", "apartment_num",
    "declared_value", "declared_value_dollar", "estimate_price", "estimate_price_dollar",
    "area_mr_bruto", "room_num2", "roof", "area_mr_neto", "floor", "warehouse",
    "first_build_year", "number_of_floors", "yard", "price_per_room", "apartments_in_building",
    "migrash", "price_per_mr", "parking", "gallery", "deal_type", "building_function",
    "house_function", "shuma_parts", "goch_appearance", "according_tva", "right_meaning",
    "neighborhood", "front", "building_pcnt", "registered_area", "building_phase_end",
    "designation", "mevune_area", "hashuma", "ground_function", "tva_detail", "front_len",
    "building_rights", "elevator_num", "field_area_mr", "scan_date",
]

NUMERIC_TARGETS = {
    "declared_profit", "sale_profit", "full_price", "declared_value",
    "declared_value_dollar", "estimate_price", "estimate_price_dollar",
    "price_per_room", "rooms_number", "room_num2", "sold_part"
}

DATE_TARGETS = {"deal_date", "sale_day"}
CHUNK_SIZE = 50000

# ---------------------------------------------------------
# Helper: Direct Excel Chunk Iterator
# ---------------------------------------------------------
def _excel_chunk_iterator(excel_path: str, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
    """
    Yields chunks of the Excel file as DataFrames directly from the stream.
    No intermediate CSV file is created.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required. pip install openpyxl")

    # read_only=True is critical for low memory usage
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    rows_iter = ws.rows
    
    # 1. Get Headers
    try:
        header_row = next(rows_iter)
        headers = [cell.value for cell in header_row]
    except StopIteration:
        wb.close()
        return

    # 2. Accumulate rows and yield chunks
    current_chunk = []
    
    for row in rows_iter:
        # Get values
        values = [cell.value for cell in row]
        current_chunk.append(values)
        
        if len(current_chunk) >= chunk_size:
            yield pd.DataFrame(current_chunk, columns=headers)
            current_chunk = []  # Reset buffer
            
    # Yield remaining
    if current_chunk:
        yield pd.DataFrame(current_chunk, columns=headers)
        
    wb.close()

# ---------------------------------------------------------
# Processing Logic (Per Chunk)
# ---------------------------------------------------------

def _process_chunk(df: pd.DataFrame) -> pd.DataFrame:
    """Apply all cleaning rules to a single chunk."""
    
    # 1. Reorder / Add missing columns
    ci_map = {str(c).lower(): c for c in df.columns}
    new_data = {}
    
    for canonical in PREFERRED_COLUMN_ORDER:
        lc = canonical.lower()
        if lc in ci_map:
            new_data[canonical] = df[ci_map[lc]]
        else:
            new_data[canonical] = pd.Series([""] * len(df), index=df.index)
            
    # Append extra columns
    used_cols = set(ci_map.values())
    mapped_cols = set([ci_map[lc] for lc in PREFERRED_COLUMN_ORDER if lc in ci_map])
    extras = [c for c in df.columns if c not in mapped_cols]
    for c in extras:
        new_data[c] = df[c]
        
    df = pd.DataFrame(new_data)
    
    # 2. Global clean
    df = df.replace("--", 0)

    # 3. Numeric Conversion
    for target in NUMERIC_TARGETS:
        if target in df.columns:
            clean_series = df[target].astype(str).str.replace(r"[^0-9\.\-]", "", regex=True)
            df[target] = pd.to_numeric(clean_series, errors="coerce")

    # 4. Full Price Calc
    if "sale_profit" in df.columns and "sold_part" in df.columns and "full_price" in df.columns:
        valid_to_calc = (
            (df["full_price"].isna()) & 
            (df["sold_part"].notna()) & 
            (df["sold_part"] != 0) & 
            (df["sale_profit"].notna())
        )
        df.loc[valid_to_calc, "full_price"] = (
            df.loc[valid_to_calc, "sale_profit"] / df.loc[valid_to_calc, "sold_part"]
        )

    # 5. Date Conversion
    for target in DATE_TARGETS:
        if target in df.columns:
            df[target] = pd.to_datetime(df[target], dayfirst=True, errors="coerce")

    # 6. Fill Logic
    if "deal_date" in df.columns and "sale_day" in df.columns:
        df["deal_date"] = df["deal_date"].fillna(df["sale_day"])
    if "city2" in df.columns and "city" in df.columns:
        df["city2"] = df["city2"].replace("", np.nan).fillna(df["city"])
    if "room_num2" in df.columns and "rooms_number" in df.columns:
        df["room_num2"] = df["room_num2"].fillna(df["rooms_number"])

    # 7. Format dates back to string
    for target in DATE_TARGETS:
        if target in df.columns:
            df[target] = df[target].dt.strftime('%d/%m/%Y')

    # 8. Text cleanup
    text_cols = df.select_dtypes(include=['object']).columns
    for col in text_cols:
        df[col] = df[col].astype(str).str.replace(",", " ", regex=False)

    # 9. Drop scan_date
    cols_lower = {str(c).lower(): c for c in df.columns}
    if "scan_date" in cols_lower:
        df = df.drop(columns=[cols_lower["scan_date"]])

    # 10. Global NaN cleanup
    df = df.replace([np.nan, "nan", "NaN", "None", "NaT"], "")
    
    return df

# ---------------------------------------------------------
# Public API
# ---------------------------------------------------------

def run_yzer_preparation(scan_path: str, output_dir: str) -> Dict[str, Any]:
    os.makedirs(output_dir, exist_ok=True)
    
    base_name = os.path.splitext(os.path.basename(scan_path))[0]
    today_str = date.today().strftime("%Y%m%d")
    output_filename = f"yzer_ready_{base_name}_{today_str}.csv"
    output_path = os.path.join(output_dir, output_filename)

    if os.path.exists(output_path):
        os.remove(output_path)

    ext = os.path.splitext(scan_path)[1].lower()
    
    # --- Select Iterator Strategy ---
    if ext in [".xls", ".xlsx", ".xlsm"]:
        # New Direct Excel Stream
        iterator = _excel_chunk_iterator(scan_path, CHUNK_SIZE)
    else:
        # CSV Stream
        encoding = "utf-8"
        for enc in ["utf-8", "cp1255", "latin1"]:
            try:
                pd.read_csv(scan_path, nrows=50, encoding=enc)
                encoding = enc
                break
            except:
                pass
        iterator = pd.read_csv(scan_path, chunksize=CHUNK_SIZE, encoding=encoding, dtype=str)

    total_rows = 0
    is_first_chunk = True

    for chunk in iterator:
        if chunk.empty: continue
        
        # 1. Process Chunk
        processed_chunk = _process_chunk(chunk)
        
        # 2. Append Chunk to File
        processed_chunk.to_csv(
            output_path, 
            mode='a', 
            index=False, 
            header=is_first_chunk, 
            encoding="utf-8-sig", 
            lineterminator='\r\n'
        )
        
        total_rows += len(processed_chunk)
        is_first_chunk = False

    return {
        "output_filename": output_filename,
        "rows_after": total_rows,
        "status": "success"
    }

def prepare_for_yzer(scan_path: str, output_dir: str):
    stats = run_yzer_preparation(scan_path, output_dir)
    return stats["output_filename"], stats["rows_after"]